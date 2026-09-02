#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
==============================================================================
MÁQUINA DE SITES — PAINEL CENTRAL DE VENDAS & ORQUESTRADOR
==============================================================================
Script principal que:
1. Executa a Prospecção com Jina AI (prospeccao.py)
2. Gera as Landing Pages com Tailwind CSS e 5 Heros (gerador_e_deploy.py)
3. Publica no Cloudflare Pages via Wrangler
4. Gera a planilha final 'leads_prontos.xlsx' contendo exatamente:
   [Nome da Empresa] | [Telefone] | [Link do Site Publicado]
5. Fornece o script de mensagem do WhatsApp pronto para copiar e enviar!
==============================================================================
"""

import os
import sys
import json
import argparse
from datetime import datetime
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr and hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')


# Importar os módulos da esteira
import prospeccao
import gerador_e_deploy

def banner():
    print("""
==============================================================================
           🚀 MÁQUINA DE SITES AUTOMÁTICA — PAINEL DE VENDAS 🚀
      Prospecção Jina AI ➔ Geração Tailwind ➔ Cloudflare Pages ➔ WhatsApp
==============================================================================
""", flush=True)

def formatar_excel(caminho_excel):
    """
    Aplica formatação visual premium na planilha leads_prontos.xlsx com openpyxl
    """
    wb = openpyxl.load_workbook(caminho_excel)
    ws = wb.active
    ws.title = "Leads Prontos"
    
    # Cores corporativas
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Slate 800
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11, color="1E293B")
    link_font = Font(name="Calibri", size=11, color="2563EB", underline="single") # Blue 600
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # 1. Estilizar Cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    # 2. Estilizar Linhas de Dados e Criar Hiperlinks Clicáveis
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        ws.row_dimensions[row_idx].height = 22
        
        # Nome da Empresa
        row[0].font = data_font
        row[0].alignment = Alignment(horizontal="left", vertical="center")
        row[0].border = thin_border
        
        # Telefone
        row[1].font = data_font
        row[1].alignment = Alignment(horizontal="center", vertical="center")
        row[1].border = thin_border
        
        # Link do Site Publicado (Hiperlink real)
        link_val = str(row[2].value or "")
        if link_val.startswith("http"):
            row[2].hyperlink = link_val
            row[2].font = link_font
        else:
            row[2].font = data_font
            
        row[2].alignment = Alignment(horizontal="left", vertical="center")
        row[2].border = thin_border

    # 3. Auto-ajustar largura das colunas
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 18)
        
    wb.save(caminho_excel)

def gerar_planilha_vendas(leads):
    """
    Gera a planilha Excel estritamente com as 3 colunas solicitadas:
    [Nome da Empresa] | [Telefone] | [Link do Site Publicado]
    """
    caminho_excel = Path("leads_prontos.xlsx")
    
    dados = []
    for lead in leads:
        dados.append({
            "Nome da Empresa": lead.get("empresa", ""),
            "Telefone": lead.get("telefone_whatsapp", lead.get("telefone", "")),
            "Link do Site Publicado": lead.get("url_publicada", "")
        })
        
    df = pd.DataFrame(dados)
    
    # Salvar em Excel
    df.to_excel(caminho_excel, index=False, engine="openpyxl")
    
    # Formatar com OpenPyXL
    formatar_excel(caminho_excel)
    
    print(f"\n📊 Planilha de Vendas Gerada: {caminho_excel.resolve()}")
    return caminho_excel

def imprimir_resumo_e_roteiro(leads):
    """
    Exibe no terminal a lista de sites e os textos de WhatsApp prontos para envio
    """
    print("\n" + "=" * 78)
    print("📋 RESUMO DE SITES GERADOS & LINKS DE WHATSAPP")
    print("=" * 78)
    
    for idx, lead in enumerate(leads, 1):
        empresa = lead.get("empresa")
        telefone = lead.get("telefone_whatsapp")
        link = lead.get("url_publicada")
        
        print(f"\n[{idx}] 🏢 {empresa.upper()}")
        print(f"    📞 WhatsApp: +{telefone}")
        print(f"    🌐 Site no Ar: {link}")
        print(f"    💬 Mensagem de Abordagem para Copiar:")
        print("    " + "-" * 70)
        
        mensagem = (
            f"Olá! Tudo bem?\n\n"
            f"Preparei uma proposta moderna para o site da *{empresa}*, já está online "
            f"para você conferir ao vivo no seu celular ou computador:\n\n"
            f"👉 {link}\n\n"
            f"O site já está otimizado para carregamento instantâneo, integrado ao seu WhatsApp e "
            f"focado em atrair novos clientes na sua região.\n\n"
            f"Podemos ajustar o que quiser! O que achou do visual?"
        )
        for linha in mensagem.split("\n"):
            print(f"    {linha}")
        print("    " + "-" * 70)

def main():
    banner()
    
    parser = argparse.ArgumentParser(description="Máquina de Sites Automática")
    parser.add_argument("--nicho", type=str, default="Odontologia", help="Nicho de prospecção (ex: Odontologia, Pizzaria, Estética)")
    parser.add_argument("--cidade", type=str, default="São Paulo, SP", help="Cidade alvo (ex: São Paulo, SP, Belo Horizonte, MG)")
    parser.add_argument("--qtd", type=int, default=5, help="Quantidade de leads a prospectar")
    parser.add_argument("--projeto", type=str, default="minha-maquina", help="Nome do projeto Cloudflare Pages")
    parser.add_argument("--jina-key", type=str, default=None, help="Chave de API Jina AI (opcional)")
    
    args = parser.parse_args()
    
    # 1. Etapa de Prospecção
    print(f"📍 ETAPA 1/3: PROSPECÇÃO COM JINA AI ({args.nicho} em {args.cidade})")
    leads = prospeccao.prospectar(
        nicho=args.nicho,
        cidade=args.cidade,
        quantidade=args.qtd,
        jina_api_key=args.jina_key or os.getenv("JINA_API_KEY")
    )
    
    if not leads:
        print("❌ Falha na prospecção. Nenhum lead encontrado.")
        sys.exit(1)
        
    # 2. Etapa de Geração e Deploy
    print(f"\n📍 ETAPA 2/3: GERAÇÃO DOS SITES & DEPLOY NO CLOUDFLARE PAGES")
    leads_publicados = gerador_e_deploy.executar_fluxo_completo(project_name=args.projeto)
    
    if not leads_publicados:
        print("❌ Falha na geração dos sites.")
        sys.exit(1)
        
    # 3. Etapa de Geração da Planilha de Vendas
    print(f"\n📍 ETAPA 3/3: EXPORTAÇÃO DA PLANILHA EXCEL DE VENDAS")
    caminho_excel = gerar_planilha_vendas(leads_publicados)
    
    # Exibir resumo e mensagens
    imprimir_resumo_e_roteiro(leads_publicados)
    
    print("\n" + "=" * 78)
    print("✨ SUCESSO! PROCESSO 100% FINALIZADO.")
    print(f"👉 Abra a planilha: {caminho_excel.name}")
    print("👉 Escolha o primeiro lead e envie a mensagem no WhatsApp!")
    print("=" * 78 + "\n")

if __name__ == "__main__":
    main()

