#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
==============================================================================
MÁQUINA DE SITES — PUBLICADOR ONLINE INSTANTÂNEO VIA CLOUDFLARE TUNNEL
==============================================================================
Cria um link HTTPS público oficial da Cloudflare (*.trycloudflare.com)
sem exigir login, sem exigir conta e sem abrir navegador para configurar!

1. Inicia um servidor HTTP local servindo a pasta /public
2. Conecta à rede global da Cloudflare via Quick Tunnel
3. Extrai a URL HTTPS pública
4. Atualiza automaticamente a planilha 'leads_prontos.xlsx'
5. Mantém o site no ar para acesso de qualquer celular no mundo!
==============================================================================
"""

import os
import sys
import re
import time
import json
import socket
import threading
import subprocess
from http.server import HTTPServer, SimpleHTTPRequestHandler
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr and hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

PORTA = 8765
PUBLIC_DIR = Path("public").resolve()

class HandlerPersonalizado(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(PUBLIC_DIR), **kwargs)
        
    def log_message(self, format, *args):
        # Silenciar logs normais para não poluir terminal
        pass

def iniciar_servidor_local(porta):
    httpd = HTTPServer(("127.0.0.1", porta), HandlerPersonalizado)
    server_thread = threading.Thread(target=httpd.serve_forever, daemon=True)
    server_thread.start()
    return httpd

def atualizar_planilha_com_url(base_url):
    caminho_json = Path("leads_publicados.json")
    caminho_excel = Path("leads_prontos.xlsx")
    
    if not caminho_json.exists():
        return []
        
    with open(caminho_json, "r", encoding="utf-8") as f:
        leads = json.load(f)
        
    for lead in leads:
        slug = lead.get("slug", "")
        lead["url_publicada"] = f"{base_url.rstrip('/')}/sites/{slug}/"
        
    with open(caminho_json, "w", encoding="utf-8") as f:
        json.dump(leads, f, ensure_ascii=False, indent=2)
        
    dados = []
    for lead in leads:
        dados.append({
            "Nome da Empresa": lead.get("empresa", ""),
            "Telefone": lead.get("telefone_whatsapp", lead.get("telefone", "")),
            "Link do Site Publicado": lead.get("url_publicada", "")
        })
        
    df = pd.DataFrame(dados)
    df.to_excel(caminho_excel, index=False, engine="openpyxl")
    
    # Formatação OpenPyXL
    wb = openpyxl.load_workbook(caminho_excel)
    ws = wb.active
    ws.title = "Leads Prontos"
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11, color="1E293B")
    link_font = Font(name="Calibri", size=11, color="2563EB", underline="single")
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        ws.row_dimensions[row_idx].height = 22
        row[0].font = data_font
        row[0].border = thin_border
        row[1].font = data_font
        row[1].alignment = Alignment(horizontal="center", vertical="center")
        row[1].border = thin_border
        link_val = str(row[2].value or "")
        if link_val.startswith("http"):
            row[2].hyperlink = link_val
            row[2].font = link_font
        row[2].border = thin_border

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 20)
        
    wb.save(caminho_excel)
    return leads

def main():
    print("""
==============================================================================
       ☁️ CLOUDFLARE QUICK TUNNEL — PUBLICADOR ONLINE INSTANTÂNEO ☁️
           Zero Configurações · Zero Logins · 100% Automático
==============================================================================
""", flush=True)

    print(f"🚀 Iniciando servidor local na pasta /public (Porta {PORTA})...", flush=True)
    iniciar_servidor_local(PORTA)
    
    print(f"⚡ Conectando à rede global da Cloudflare...", flush=True)
    cmd = f"npx cloudflared tunnel --url http://localhost:{PORTA}"
    
    process = subprocess.Popen(
        cmd,
        shell=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        encoding="utf-8",
        errors="replace"
    )
    
    url_publica = None
    print("⏳ Aguardando geração da URL pública segura pela Cloudflare...\n", flush=True)
    
    while True:
        line = process.stdout.readline()
        if not line and process.poll() is not None:
            break
            
        if line:
            # Detectar linha da URL
            match = re.search(r'(https://[a-zA-Z0-9\-_.]+\.trycloudflare\.com)', line)
            if match and not url_publica:
                url_publica = match.group(1)
                print("=" * 78, flush=True)
                print("🎉 SUCESSO ABSOLUTO! SEUS SITES ESTÃO NO AR NA CLOUDFLARE!", flush=True)
                print(f"🌐 DOMÍNIO PÚBLICO ATIVO: {url_publica}", flush=True)
                print("=" * 78, flush=True)
                
                # Atualizar planilha
                leads = atualizar_planilha_com_url(url_publica)
                print(f"\n📊 Planilha 'leads_prontos.xlsx' atualizada com os links reais!", flush=True)
                print("\n📱 LINKS REAIS PRONTOS PARA ACESSAR NO CELULAR OU WHATSAPP:\n", flush=True)
                
                for idx, lead in enumerate(leads, 1):
                    empresa = lead.get("empresa", "")
                    link = lead.get("url_publicada", "")
                    print(f"  [{idx}] {empresa}:", flush=True)
                    print(f"      👉 {link}\n", flush=True)
                    
                print("=" * 78, flush=True)
                print("🟢 O servidor está ativo e respondendo aos acessos de qualquer lugar!", flush=True)
                print("=" * 78, flush=True)
                
        time.sleep(0.05)

if __name__ == "__main__":
    main()

